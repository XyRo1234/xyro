from tkinter import * # __all__
from tkinter import filedialog
import tkinter as tk
from tkinter import PhotoImage, Button, Label, StringVar, ttk, colorchooser
import tkinter.ttk as ttk
import tkinter.messagebox as msgbox
import os
import pandas as pd
from pyairtable import Table
import re
import numpy
from chardet import detect
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
import datetime

'''test23_엑셀toSRT_전환기능_2 / test23_2D텍스트뽑기 두개합친파일'''
'''Airtable에서 다운받은 csv파일로 srt, txt, 2Dtext엑셀파일 생성'''

root = Tk()
root.title("Airtable Excel Transform")
#root.geometry("640x480+300+100") # 가로 * 세로 + x좌표 + y좌표

dnow = datetime.datetime.now()


# 파일선택 추가
def add_file():
    files = filedialog.askopenfilenames(title="이미지 파일을 선택하세요", \
        filetypes=(("엑셀 파일", "*.csv *.xlsx"), ("모든 파일", "*.*")), \
        initialdir=r"D:\Program Files\workspace_2\1_Airtable_script")
        # 최초에 사용자가 지정한 경로를 보여줌
    
    # 사용자가 선택한 파일 목록
    for file in files:
        list_file.insert(END, file)

# 선택 삭제
def del_file():
    #print(list_file.curselection())
    for index in reversed(list_file.curselection()):
        list_file.delete(index)



# 저장 경로 (폴더)
def browse_dest_path():
    folder_selected = filedialog.askdirectory()
    if folder_selected == '': # 사용자가 취소를 누를 때
        return
    #print(folder_selected)
    txt_dest_path.delete(0, END)
    txt_dest_path.insert(0, folder_selected)

def start_1():
    if list_file.size() == 0:
        msgbox.showwarning("경고", "파일을 선택하세요")
        return
    for idx, excelfile in enumerate(list_file.get(0, END)):
        print(excelfile)
        path = txt_dest_path.get()

        # 파일불러오기 / dataframe만들기

        if '.csv' in excelfile:
            try:
                df = pd.read_csv(f'{excelfile}', encoding='CP949',  na_values=["?", "??", "N/A", ""], keep_default_na= False)  # 엑셀값 dataframe만들기
            except UnicodeDecodeError:
                df = pd.read_csv(f'{excelfile}', encoding='utf-8',  na_values=["?", "??", "N/A", ""], keep_default_na= False)  # 엑셀값 dataframe만들기
        else:
            df = pd.read_excel(f'{excelfile}')

        header = df.columns.values.tolist()     # BOM값으로 인해 첫열 이름이 이상하게 변형되었을때, 열이름 변경하기
        if '癤풬o' in header:                   # BOM값으로 인해 첫열 이름이 이상하게 변형되었을때, 열이름 변경하기
            df.rename(columns={'癤풬o':'Ep_No'}, inplace=True) # 열이름 변경하기

        df = df.filter(items=['No','Episode','SRT_Time_Code','SRT/SSML'])     # df필요한 열만 추출
        df.dropna(subset=['Episode'], axis=0, inplace=True)                   # Episode열의 결측치(nan) 제거

        df_ep_list_draft = df['Episode'].values.tolist()    # Episode 리스트 만들기 ( [E01,E01,E02,E02,E02,E03,...] )
        df_ep_list = []                             # Episode 리스트 중복값 제거
        for i,value in enumerate(df_ep_list_draft): # Episode 리스트 중복값 제거
            if value not in df_ep_list:             # Episode 리스트 중복값 제거
                df_ep_list.append(value)            # Episode 리스트 중복값 제거
        df_ep_list.sort
        print(df_ep_list)

        for i,value in enumerate(df_ep_list):           # Episode별 순차진행
            df_ep = df.loc[(df['Episode'] == value)]    # Episode별 순차진행/특정Ep의 df만 남기기
            # print(df_ep)
            # 파일생성 및 data 입력
            with open(f'{path}\\(SRT){value}_{dnow.strftime("%y%m%d")}.srt', 'w', encoding='UTF-8-SIG') as fl:   # 파일 생성
                for i in range(0,df_ep.shape[0]):      # df.shape = [행수,열수] / 따라서 행수만큼 반복
                    fl.write('\n')
                    fl.write(str(i+1))
                    fl.write('\n')
                    fl.write(str(df_ep.loc[(df_ep['No']==i+1)]['SRT_Time_Code'].values[0]).strip())
                    fl.write('\n')
                    fl.write(str(df_ep.loc[(df_ep['No']==i+1)]['SRT/SSML'].values[0]).strip())
                    # print(str(df_ep.loc[(df_ep['No']==i+1)]['SRT/SSML'].values[0]).strip())
                    fl.write('\n')

            with open(f'{path}\\{value}_{dnow.strftime("%y%m%d")}.txt', 'w', encoding='UTF-8-SIG') as fl:   # 파일 생성
                for i in range(0,df_ep.shape[0]):      # df.shape = [행수,열수] / 따라서 행수만큼 반복
                    fl.write(str(df_ep.loc[(df_ep['No']==i+1)]['SRT/SSML'].values[0]).strip())
                    # print(str(df_ep.loc[(df_ep['No']==i+1)]['SRT/SSML'].values[0]).strip())
                    fl.write('\n')

            print(value,'완료')

            progress = (i + 1) / len(df_ep_list) * 100 # 실제 percent 정보를 계산
            p_var.set(progress)


def start_2():
    if list_file.size() == 0:
        msgbox.showwarning("경고", "파일을 선택하세요")
        return
    #파일 불러오기
    for i,excelfile in enumerate(list_file.get(0, END)):
        path = txt_dest_path.get()
        if '.csv' in excelfile:
            try:
                df = pd.read_csv(f'{excelfile}', encoding='CP949',  na_values=["?", "??", "N/A", ""], keep_default_na= False)  # 엑셀값 dataframe만들기
            except UnicodeDecodeError:
                df = pd.read_csv(f'{excelfile}', encoding='utf-8',  na_values=["?", "??", "N/A", ""], keep_default_na= False)  # 엑셀값 dataframe만들기
        else:
            df = pd.read_excel(f'{excelfile}')
        print(i)
        print(excelfile)

        header = df.columns.values.tolist()     # BOM값으로 인해 첫열 이름이 이상하게 변형되었을때, 열이름 변경하기
        if '癤풬o' in header:                   # BOM값으로 인해 첫열 이름이 이상하게 변형되었을때, 열이름 변경하기
            df.rename(columns={'癤풬o':'Ep_No'}, inplace=True) # 열이름 변경하기
        df = df.filter(items=['2D Text'])     # df필요한 열만 추출

        p = re.compile("\d{2}")
        m = p.search(excelfile)
        newfile_name = 'E'+m.group() # E01 등

        # xlsx_name = excelfile.split('.csv')[0] + '.xlsx'    # 파일명 확장자 csv -> xlsx로 바꾸기
        xlsx_name = path + '\\2D Text_'+dnow.strftime("%y%m%d")+'.xlsx'       # 생성되는 2D Text 엑셀파일명

        """ 엑셀파일생성 및 첫 시트 만들기 """
        if excelfile == list_file.get(0, END)[0]:                # 첫번째 파일이면 실행해라!
            with pd.ExcelWriter(f'{xlsx_name}', mode='w', engine='openpyxl') as writer: # 엑셀생성
                df.to_excel(writer,sheet_name=newfile_name, index=False)                # 엑셀생성
        else:                                         # 첫번째 파일이 아니면 실행해라!
            # wsname = file[5:8]                      # 파일의 5:8 index값을 name으로 지정
            with pd.ExcelWriter(f'{xlsx_name}', mode='a', engine='openpyxl') as writer: # 엑셀시트추가
                df.to_excel(writer,sheet_name=newfile_name, index=False)                # 엑셀시트추가
        
    wb = load_workbook(xlsx_name)
    sheet = wb.active
    wsnames = wb.get_sheet_names()      # 모든 시트명 가져오기
    for i,wsname in enumerate(wsnames):
        sheet = wb[wsname]              # 해당시트명의 시트 열기
        sheet.column_dimensions["A"].width = 70
        sheet.column_dimensions["B"].width = 70
        sheet.column_dimensions["C"].width = 70
        sheet.delete_rows(1)
        for x in range(1, sheet.max_row + 1):
            sheet[f'A{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
            sheet[f'B{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
            sheet[f'C{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
        wb.save(xlsx_name)



# Column = 0 사이즈 조절                                                                   # 0,0
column_frame = Label(root, text="", width=60, height=0)
column_frame.grid(row=0, column=0)


# 파일 프레임 (파일 추가, 선택 삭제)                                                        # 1,0
file_frame = LabelFrame(root, text="파일선택")
file_frame.grid(row=1, column=0, padx=5, pady=5, sticky=N+E+W+S)

btn_add_file = Button(file_frame, padx=5, pady=5, width=12, text="파일추가", command=add_file)
btn_add_file.pack(side="left")

btn_del_file = Button(file_frame, padx=5, pady=5, width=12, text="선택삭제", command=del_file)
btn_del_file.pack(side="right")

# 리스트 프레임                                                                            # 2,0
list_frame = Frame(root)
list_frame.grid(row=2, column=0, padx=5, pady=5, sticky=E+W)

scrollbar = Scrollbar(list_frame)
scrollbar.pack(side="right", fill="y")

list_file = Listbox(list_frame, selectmode="extended", height=15, yscrollcommand=scrollbar.set)
list_file.pack(side="left", fill="both", expand=True)
scrollbar.config(command=list_file.yview)

# 저장 경로 프레임                                                                         # 3,0
path_frame = LabelFrame(root, text="저장경로")
path_frame.grid(row=3, column=0, padx=5, pady=5, sticky=N+E+W+S)


txt_dest_path = Entry(path_frame)
txt_dest_path.pack(side="left", fill="x", expand=True, padx=5, pady=5, ipady=4) # 높이 변경

btn_dest_path = Button(path_frame, text="저장경로", width=10, command=browse_dest_path)
btn_dest_path.pack(side="right", padx=5, pady=5)



# 진행 상황 Progress Bar                                                                   # 4,0
frame_progress = LabelFrame(root, text="진행상황")
frame_progress.grid(row=4, column=0, padx=5, pady=20, sticky=E+W)

p_var = DoubleVar()
progress_bar = ttk.Progressbar(frame_progress, maximum=100, variable=p_var)
progress_bar.pack(fill="x", padx=5, pady=5)


# 실행 프레임                                                                              # 5,0
frame_run = Frame(root)
frame_run.grid(row=5, column=0)

btn_close = Button(frame_run, padx=5, pady=5, text="닫기", width=12, command=root.quit)
btn_close.pack(side="right", padx=5, pady=5)

btn_start = Button(frame_run, padx=5, pady=5, text="SRT 생성", width=12, command=start_1)
btn_start.pack(side="right", padx=5, pady=5)


# Column = 0 사이즈 조절                                                                   # 0,2
column_frame = Label(root, text="정원석", width=60, height=0)
column_frame.grid(row=0, column=2)

# Column = 0 사이즈 조절                                                                   # 0,2
column_frame = Label(root, text="2D Text 엑셀생성", width=60, height=0)
column_frame.grid(row=4, column=2)

# 실행 프레임                                                                              # 5,2
frame_run = Frame(root)
frame_run.grid(row=5, column=2)

btn_close = Button(frame_run, padx=5, pady=5, text="닫기", width=12, command=root.quit)
btn_close.pack(side="right", padx=5, pady=5)

btn_start = Button(frame_run, padx=5, pady=5, text="2D Text 생성", width=12, command=start_2)
btn_start.pack(side="right", padx=5, pady=5)


# Column = 0 하단여백                                                                      # 100,0
column_frame = Label(root, text="", width=0, height=0)
column_frame.grid(row=100, column=0)



root.resizable(True, False) # x(너비), y(높이) 값 변경 불가 (창 크기 변경 불가)
root.mainloop()