from tkinter import * # __all__
from tkinter import filedialog
import tkinter as tk
from tkinter import PhotoImage, Button, Label, StringVar, ttk, colorchooser
import tkinter.ttk as ttk
import tkinter.messagebox as msgbox
import os
import pandas as pd
from pyairtable import Table
import re
import numpy
from chardet import detect
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
import datetime

dnow = datetime.datetime.now()

excelfile ='D:\\Program Files\\Workspace\\airtable\\VF 스크립트.xlsx'
path = 'D:\\Program Files\\Workspace\\airtable'

'''
[Time Code]  [Modified Script]	[2D Text]	[Navigation Bar]	[참고 Link]	[장면]	[장면 설명]	[Unreal]	[UR 소품]	[AE]	[AE 소품]
'''


def start_1(sheet_name, path, df):
    ''' SRT 추출 '''
    df_script = df.filter(items=['Time Code','Modified Script'])   # df필요한 열만 추출
    df_script = df.dropna(subset=['Modified Script'])       # 해당 열에서 빈칸이 있는 행 삭제
    df_script = df_script.reset_index(drop=True)            # 행값 리셋 0,1,2,3, ...
    # df_script = df_script.apply

    with open(f'{path}\\(SRT){sheet_name}_{dnow.strftime("%y%m%d")}.srt', 'w', encoding='UTF-8-SIG') as fl:   # 파일 생성
        for i in range(0,df_script.shape[0]):      # df.shape = [행수,열수] / 따라서 행수만큼 반복
            # fl.write('\n')
            fl.write(str(i+1))
            fl.write('\n')
            try:
                if str(df_script.loc[i]['Time Code']).strip() == 'nan':
                    pass
                else:
                    fl.write(str(df_script.loc[i]['Time Code']).strip())
            except KeyError:
                pass
            fl.write('\n')
            fl.write(str(df_script.loc[i,'Modified Script']).strip())
            # print(str(df_script.loc[(df_script['No']==i+1)]['SRT/SSML'].values[0]).strip())
            fl.write('\n')
            fl.write('\n')

    '''txt추출'''
    with open(f'{path}\\{sheet_name}_{dnow.strftime("%y%m%d")}.txt', 'w', encoding='UTF-8-SIG') as fl:   # 파일 생성
        for i in range(0,df_script.shape[0]):      # df.shape = [행수,열수] / 따라서 행수만큼 반복
            fl.write(str(df_script.loc[i,'Modified Script']).strip())
            # print(str(df_script.loc[(df_script['No']==i+1)]['SRT/SSML'].values[0]).strip())
            fl.write('\n')

'''
sheet 반복함수 // SRT 추출
'''

wb = load_workbook(excelfile)
sheet = wb.active
wsnames = wb.get_sheet_names()      # 모든 시트명 가져오기

sheet_name_list = []    # [E01_How to A, E02_How to B, ...]
episode_list = []       # [E01, E02, E03, ...]
for sheet_name in wsnames:
    m = re.compile("E\d{2}").search(sheet_name)
    if m:
        sheet_name_list.append(sheet_name)
        episode_list.append(m.group())
print(sheet_name_list)
print(episode_list)
wb.close()

for sheet_name in sheet_name_list:  # SRT / txt 추출
    df = pd.read_excel(f'{excelfile}', sheet_name=sheet_name)
    print('SRT생성 ' + sheet_name)
    start_1(sheet_name, path, df)

'''
2D Text 추출 : 엑셀생성/시트별 데이터 삽입
'''
xlsx_name = path + '\\Text_All_'+dnow.strftime("%y%m%d")+'.xlsx'       # 생성되는 2D Text 엑셀파일명
for i, sheet_name in enumerate(sheet_name_list):
    df = pd.read_excel(f'{excelfile}', sheet_name=sheet_name)
    df_2d = df.filter(items=['2D Text'])     # df필요한 열만 추출
    df_2d = df_2d.dropna(subset=['2D Text'])       # 해당 열에서 빈칸이 있는 행 삭제
    df_2d = df_2d['2D Text'].astype(str).str.strip()

    """ 엑셀파일생성 및 첫 시트 만들기 """
    if sheet_name == sheet_name_list[0]:                # 첫번째 시트면 실행해라!
        with pd.ExcelWriter(f'{xlsx_name}', mode='w', engine='openpyxl') as writer: # 엑셀생성
            df_2d.to_excel(writer,sheet_name=episode_list[0], index=False)                # 엑셀생성
        print('2D Text 시트생성 ' + episode_list[0])
    else:                                         # 첫번째 시트가 아니면 실행해라!
        # wsname = file[5:8]                      # 파일의 5:8 index값을 name으로 지정
        with pd.ExcelWriter(f'{xlsx_name}', mode='a', engine='openpyxl') as writer: # 엑셀시트추가
            df_2d.to_excel(writer,sheet_name=episode_list[i], index=False)                # 엑셀시트추가
        print('2D Text 시트생성 ' + episode_list[i])

'''
2D Text 추출 : 열간격 및 왼쪽센터정렬
'''
wb = load_workbook(xlsx_name)
sheet = wb.active
wsnames = wb.get_sheet_names()      # 모든 시트명 가져오기
for i,wsname in enumerate(wsnames):
    sheet = wb[wsname]              # 해당시트명의 시트 열기
    sheet.column_dimensions["A"].width = 70
    sheet.column_dimensions["B"].width = 70
    sheet.column_dimensions["C"].width = 70
    sheet.delete_rows(1)   
    for x in range(1, sheet.max_row + 1):
        sheet[f'A{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
        sheet[f'B{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
        sheet[f'C{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
    wb.save(xlsx_name)
    print('열간격설정 '+wsname)
wb.close()


# df_script.to_excel(f'{path}\\test.xlsx')