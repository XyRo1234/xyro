import os
import pandas as pd
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
import re

path = 'D:\\Program Files\\Workspace\\00_Create\\'
path2 = 'D:\\Program Files\\Workspace\\00_Default\\UK_1\\'    # 참고할 영어 SRT파일
path3 = 'D:\\Program Files\\Workspace\\00_Default\\UK_SRTs\\' # 참고할 다국어 SRT파일
excel_filenames_draft = os.listdir(path)
srt_filenames_draft = os.listdir(path2)


""" 엑셀 및 영어srt파일 리스트만들기 """
excel_filenames = []                    # 골라낸 파일넣을 리스트 생성
for i in excel_filenames_draft:
    if '.xlsx' in i:              # 파일확장자 리스트 골라내기
        excel_filenames.append(i)
excel_filenames.sort()                  # 리스트 abc순 정렬
srt_filenames = []                    # 골라낸 파일넣을 리스트 생성
for i in srt_filenames_draft:
    if '.srt' in i:              # 파일확장자 리스트 골라내기
        srt_filenames.append(i)
srt_filenames.sort()                  # 리스트 abc순 정렬


""" 다국어srt파일 path+filename으로 리스트만들기 """
CC_list = []
for (AR_path, dir, files) in os.walk(path3):
    for filename in files:
        ext = os.path.splitext(filename)[-1]
        if ext == '.srt':
            CC_list.append(AR_path+"\\"+ filename)


p = re.compile("E\d{2}")
p2 = re.compile("..-..")


for filename in excel_filenames:              # 엑셀파일(각 국가 피드백엑셀파일) 열기
    wb = load_workbook(f'{path}{filename}')
    sheet = wb.active
    wsnames = wb.get_sheet_names()      # 모든 시트명 가져오기
    AR_filenames = []
    # print(m2.group())
    
    for AR_pathfilename in CC_list:                     # 매치(ex fr-FR)되는 다국어 srt파일명찾기
        if p2.search(filename).group() in AR_pathfilename:               # 매치(ex fr-FR)되는 다국어 srt파일명찾기
            AR_filenames.append(AR_pathfilename)        # 매치(ex fr-FR)되는 다국어 srt파일명찾기
    
    # print(AR_filenames)

    for i,wsname in enumerate(wsnames):
        if p.search(wsname):
            sheet = wb[wsname]              # 해당시트명의 시트 열기
            # sheet.insert_cols(5, 2)         # C뒤 2열추가
            # sheet.insert_cols(3, 2)         # B뒤 2열추가
            # sheet.column_dimensions["C"].width = 10
            # sheet.column_dimensions["D"].width = 10
            # sheet.column_dimensions["G"].width = 10
            # sheet.column_dimensions["H"].width = 10
            m = p.search(wsname)

            for a in srt_filenames:            # 매치(E00)되는 영어 srt파일명찾기
                if m.group() in a:             # 매치(E00)되는 영어 srt파일명찾기
                    En_srt_filename = a        # 매치(E00)되는 영어 srt파일명찾기
            print(En_srt_filename)    
            for b in AR_filenames:             # 매치(E00)되는 다국어 srt파일명찾기
                if m.group() in b:             # 매치(E00)되는 다국어 srt파일명찾기
                    AR_srt_filename = b        # 매치(E00)되는 다국어 srt파일명찾기
            print(AR_srt_filename)
            with open(f"{path2}{En_srt_filename}", 'r', encoding="utf-8") as fr:   # 매치되는 영어 srt파일 열기
                En_lines = fr.readlines()            
            with open(AR_srt_filename, 'r', encoding="utf-8") as fr:   # 매치되는 다국어 srt파일 열기
                AR_lines = fr.readlines()

                """ 2D """
            for k in range(6, 20):                      # 엑셀시트 열수k, 검색제한 6~19열까지만
                for i,AR_line_data in enumerate(AR_lines):    # 다국어srt 라인 훓기
                    if k >= 6 :
                        C_data = str(sheet[f'C{k+1}'].value).strip()    # 열수정 주의!  
                        # print(C_data.ljust(20),AR_line_data)
                        if C_data in AR_line_data:
                            SRT_line_number = AR_lines[int(i-2)]
                            # print(SRT_line_number + '  SRT라인넘버')
                            sheet[f'A{k+1}'] = SRT_line_number
                            # for q, En_line_data in enumerate(En_lines):
                                # if SRT_line_number == En_line_data:
                                    # print(En_lines[q] + "  En프린트")
                                    # print(En_lines[q+2])
                                    # sheet[f'B{k+1}'] = En_lines[q+2]
                        else:
                            continue

                """ 자막 """
            for k in range(6, 20):                      # 엑셀시트 열수k, 검색제한 6~19열까지만
                for i,AR_line_data in enumerate(AR_lines):    # 다국어srt 라인 훓기
                    if k >= 6 :
                        C_data = str(sheet[f'C{k+1}'].value).strip()    # 열수정 주의!  
                        # print(C_data.ljust(20),AR_line_data)
                        if C_data in AR_line_data:
                            SRT_line_number = AR_lines[int(i-2)]
                            # print(SRT_line_number + '  SRT라인넘버')
                            sheet[f'I{k+1}'] = SRT_line_number
                            for q, En_line_data in enumerate(En_lines):
                                if SRT_line_number == En_line_data:
                                    # print(En_lines[q] + "  En프린트")
                                    # print(En_lines[q+2])
                                    sheet[f'J{k+1}'] = En_lines[q+2]
                        else:
                            continue
        else:
            print("해당시트는 패스합니다. " + wsname)
    wb.save(f'{path}{filename}')
