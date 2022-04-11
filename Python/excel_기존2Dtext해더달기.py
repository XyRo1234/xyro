import os
import pandas as pd
from pandas import ExcelWriter
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook

path = 'D:\\Program Files\\Workspace\\others\\'
filenames_draft = os.listdir(path)

filenames = []                    # 골라낸 파일넣을 리스트 생성
for i in filenames_draft:
    if '.xlsx' in i:              # 파일확장자 리스트 골라내기
        filenames.append(i)
filenames.sort()                  # 리스트 abc순 정렬

for file in filenames:
    wb = load_workbook(f"{path}{file}")
    ws = wb.active
    wsnames = wb.get_sheet_names()      # 모든 시트명 가져오기
    for i,wsname in enumerate(wsnames):
        sheet = wb[wsname]              # 해당시트명의 시트 열기
        sheet.insert_rows(1)            # 한줄 내리기
        sheet['A1'] = 'en-Origin'
        sheet['A1'].font = Font(bold=True)
        sheet['B1'] = 'en-Modify'
        sheet['B1'].font = Font(bold=True)
        sheet['C1'] = 'fr-CA'
        sheet['C1'].font = Font(bold=True)
        sheet.column_dimensions["A"].width = 50
        sheet.column_dimensions["B"].width = 50
        sheet.column_dimensions["C"].width = 50
        for x in range(1, sheet.max_row + 1):
            sheet[f'A{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
            sheet[f'B{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
            sheet[f'C{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
    
    wb.save(f"{path}{file}")