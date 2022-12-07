import os
import pandas as pd
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
import re

file = 'D:\\Program Files\\Workspace\\00_Default\\BF_VB_2D Text_(bg-BG)_221019.xlsx'


# 파일오픈
wb = load_workbook(file)
wsnames_1 = wb.get_sheet_names()      # 모든 시트명 가져오기
for index_1, wsname_1 in enumerate(wsnames_1):
    print('////////// '+wsname_1+' //////////')
    sheet = wb[wsname_1]              # 해당시트명의 시트 열기
    ## data 가져오기 ##
    for x in range(1, sheet.max_row + 1):
        # print(f'{sheet.cell(row=x, column=1).value:<50}','  ',f'{sheet.cell(row=x, column=2).value:<50}')
        data1 = sheet.cell(row=x, column=1).value.strip()
        data2 = sheet.cell(row=x, column=2).value.strip()
        if data1 == data2:
            fill = PatternFill("solid", start_color="00FFFF00", end_color="00FFFF00")
            sheet.cell(row=x, column=2).fill = fill

        # # print(sheet2.cell(row=x, column=1).value)
        # data = sheet2.cell(row=x, column=1).value
        # match = f"=MATCH(A{x},B{x},0)"
        # sheet.cell(row=x, column=2, value=data)
        # sheet.cell(row=x, column=3, value=match)

wb.save(file)