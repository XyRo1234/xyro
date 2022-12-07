import pandas as pd
import os
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
import numpy as np
import re

"""
각각의 EP의 영어원본txt, 영어수정본txt, 프랑스어번역본 / 3개의 파일을 각각의 EP의 엑셀시트로 통합.
"""

Create_path = 'D:\\Program Files\\Workspace\\00_Create\\'
en_path = 'D:\\Program Files\\Workspace\\00_Default\\'
bb_path = 'D:\\Program Files\\Workspace\\00_Default_2\\'
cc_path = 'D:\\Program Files\\Workspace\\00_Default_3\\'

en_filenames = os.listdir(en_path)
bb_filenames = os.listdir(bb_path)
cc_filenames = os.listdir(cc_path)

en_txt_filenames = []                    # 골라낸 파일 리스트
for i in en_filenames:
    if '.txt' in i:              # 'txt' or 'srt'가 포함되어있는 리스트 골라내기
        en_txt_filenames.append(i)
en_txt_filenames.sort()

bb_txt_filenames = []
for i in bb_filenames:
    if '.txt' in i:              # 'txt' or 'srt'가 포함되어있는 리스트 골라내기
        bb_txt_filenames.append(i)
bb_txt_filenames.sort()

cc_txt_filenames = []
for i in cc_filenames:
    if '.txt' in i:              # 'txt' or 'srt'가 포함되어있는 리스트 골라내기
        cc_txt_filenames.append(i)
cc_txt_filenames.sort()


def new_func(Create_path, newfile_name):
    """뽑은엑셀에서 열너비 지정, 오른쪽가운데정렬, 자동줄바꿈 활성화 시켜서 저장"""
    wb = load_workbook(f"{Create_path}test_{newfile_name}.xlsx")
    return wb

for n, en_file in enumerate(en_txt_filenames):

    en_line_list = []
    with open(f'{en_path}{en_txt_filenames[n]}','r',encoding = 'utf8') as File_Source:
        data = File_Source.readlines()
        for l in data:
            en_line_list.append(l.strip())

    bb_line_list = []
    try:
        with open(f'{bb_path}{bb_txt_filenames[n]}','r',encoding = 'utf8') as File_Source:
            data = File_Source.readlines()
            for l in data:
                bb_line_list.append(l.strip())
    except IndexError:
        print('bb_line_list IndexError'+ str(en_file))


    cc_line_list = []
    try:
        with open(f'{cc_path}{cc_txt_filenames[n]}','r',encoding = 'utf8') as File_Source:
            data = File_Source.readlines()
            for l in data:
                cc_line_list.append(l.strip())
    except IndexError:
        print('cc_line_list IndexError'+ str(en_file))


    dic = {
        'en-Origin': en_line_list,
        'en-Modify': bb_line_list,
        'fr-CA': cc_line_list
    }

    df = pd.DataFrame.from_dict(dic, orient='index')
    df = df.transpose()     # DataFrame 행열 전환

    p = re.compile("E\d{2}")
    m = p.search(en_file)
    newfile_name = m.group()

    """ DataFrame을 엑셀로 뽑기 """
    df.to_excel(f'{Create_path}test_{newfile_name}.xlsx',sheet_name=f'{newfile_name}', header = True, index = False)

    wb = new_func(Create_path, newfile_name)
    ws = wb.active
    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 70
    for x in range(1, ws.max_row + 1):
    # for y in range(1, ws.max_column + 1):
        ws[f'A{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
        ws[f'B{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
        ws[f'C{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
    wb.save(f'{Create_path}test_{newfile_name}.xlsx')