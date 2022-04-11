import os
import pandas as pd
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
import re

"""
EP01, EP02, ... 생성된 엑셀파일을 하나의 엑셀파일에 통합
"""

path = 'D:\\Program Files\\Workspace\\00_Create\\'
final_name = 'Script_sheets.xlsx'
filenames_draft = os.listdir(path)

filenames = []                    # 골라낸 파일넣을 리스트 생성
for i in filenames_draft:
    if '.xlsx' in i:              # 파일확장자 리스트 골라내기
        filenames.append(i)
filenames.sort()                  # 리스트 abc순 정렬



""" 엑셀파일생성 및 첫 시트 만들기 """
df = pd.read_excel(f'{path}{filenames[0]}')

p = re.compile("E\d{2}")
m = p.search(filenames[0])
wsname = m.group()

# wsname = (filenames[0])[5:8]
with pd.ExcelWriter(f'{path}{final_name}', mode='w', engine='openpyxl') as writer:
    df.to_excel(writer,sheet_name=wsname, index=False)


""" 두번째시트부터 추가하기 """
for file in filenames:
    p = re.compile("E\d{2}")
    m = p.search(file)
    wsname = m.group()
    if file != filenames[0]:                    # 첫번째 파일이 아니면 실행해라!
        df = pd.read_excel(f'{path}{file}')     # df생성
        # wsname = file[5:8]                      # 파일의 5:8 index값을 name으로 지정
        with pd.ExcelWriter(f'{path}{final_name}', mode='a', engine='openpyxl') as writer:
            df.to_excel(writer,sheet_name=wsname, index=False)


wb = load_workbook(f'{path}{final_name}')
sheet = wb.active
wsnames = wb.get_sheet_names()      # 모든 시트명 가져오기
for i,wsname in enumerate(wsnames):
    sheet = wb[wsname]              # 해당시트명의 시트 열기
    hearder=[]
    for seq, name in enumerate(range(1, sheet.max_column + 1)):
        header = sheet.cell(row=1, column=seq+1)
        header.font = Font(color="000000FF", bold=True, size = 14)
        header.alignment = Alignment(horizontal="center", vertical="center")
    # sheet['A1'] = 'en-Origin'
    # sheet['A1'].font = Font(bold=True)
    # sheet['B1'] = 'en-Modity'
    # sheet['B1'].font = Font(bold=True)
    # sheet['C1'] = 'fr-CA'
    # sheet['C1'].font = Font(bold=True)
    sheet.column_dimensions["A"].width = 70
    sheet.column_dimensions["B"].width = 70
    sheet.column_dimensions["C"].width = 70
    for x in range(1, sheet.max_row + 1):
        sheet[f'A{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
        sheet[f'B{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
        sheet[f'C{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
    wb.save(f'{path}{final_name}')