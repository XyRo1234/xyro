import os
import glob
import re
import os.path
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
import datetime

dnow = datetime.datetime.now()
path = "D:\\영상콘텐츠\\2_VB\\srt"


def a1(path_lang,path_excel,path_en):
    # path_lang = "D:\\Program Files\\Workspace\\00_Default\\"
    # path_excel = "D:\\Program Files\\Workspace\\00_Create\\"
    # path_en = "D:\\Program Files\\Workspace\\00_Original\\"

    ###### <파일명 확장자 변경> #############################
    files_lang = glob.glob(f'{path_lang}*.*')
    files_en = glob.glob(f'{path_en}*.*')

    for name in files_lang:

        if not os.path.isdir(name):  # 디렉토리는 포함 X
            filename1 = os.path.splitext(name)  # 확장자와 파일명 구분
            # print('filename1:',filename1[0],"//", filename1[1]) # 확장자와 파일명 구분 확인
            try:
                os.rename(name, filename1[0] + '.txt')  # ssml을 없앤 파일명에 txt 확장자 붙이기
            except:
                pass

    for name in files_en:

        if not os.path.isdir(name):  # 디렉토리는 포함 X
            filename1_original = os.path.splitext(name)  # 확장자와 파일명 구분
            # print('filename1:',filename1[0],"//", filename1[1]) # 확장자와 파일명 구분 확인
            try:
                # ssml을 없앤 파일명에 txt 확장자 붙이기
                os.rename(name, filename1_original[0] + '.txt')
            except:
                pass
    ########################################################


    ##### <TXT파일로 확장자 변경후 파일 list 만들기> #########
    filename_lang = []                  # 파일명 리스트
    filename_lang_draft = os.listdir(path_lang)
    for i in filename_lang_draft:
        if '.txt' in i:              # 'txt' 가 포함되어있는 리스트 골라내기
            filename_lang.append(i)

    filename_en = []
    filename_en_draft = os.listdir(path_en)
    for o in filename_en_draft:
        if '.txt' in o:
            filename_en.append(o)


    ''' 엑셀파일 생성 '''

    for n, filename in enumerate(filename_lang):

        en_line_list = []
        with open(f'{path_en}{filename_en[n]}','r',encoding = 'utf8') as File_Source:
            data = File_Source.readlines()
            data = list(map(lambda s: s.strip(), data))
            for l in data:
                en_line_list.append(l.strip())

        bb_line_list = []
        with open(f'{path_lang}\\{filename_lang[n]}','r',encoding = 'utf8') as File_Source:
            data = File_Source.readlines()
            data = list(map(lambda s: s.strip(), data))
            for l in data:
                bb_line_list.append(l.strip())

        dic = {
            'en': en_line_list,
            'Target language': bb_line_list,
            'Feedback': ''
        }

        df = pd.DataFrame.from_dict(dic, orient='index')
        df = df.transpose()     # DataFrame 행열 전환


        p = re.compile("E\d{2}")
        m = p.search(filename)
        episode = m.group()
        newfile_name = os.path.splitext(filename)[0]   # 확장자와 파일명 구분

        """ DataFrame을 엑셀로 뽑기 """
        df.to_excel(f'{path_excel}\\{newfile_name}.xlsx',sheet_name=f'{episode}', header = True, index = False)

        wb = load_workbook(f'{path_excel}\\{newfile_name}.xlsx')
        ws = wb.active
        ws.column_dimensions["A"].width = 70
        ws.column_dimensions["B"].width = 70
        ws.column_dimensions["C"].width = 70
        for x in range(1, ws.max_row + 1):
        # for y in range(1, ws.max_column + 1):
            ws[f'A{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
            ws[f'B{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
            ws[f'C{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
        wb.save(f'{path_excel}\\{newfile_name}.xlsx')
        print(episode,'완료')


def a2(path,lang):
    """
    EP01, EP02, ... 생성된 엑셀파일을 하나의 엑셀파일에 통합
    """
    # path = 'D:\\Program Files\\Workspace\\00_Create\\'
    final_name = '1_Script_sheets_('+lang+')_'+dnow.strftime("%y%m%d")+'.xlsx'
    filenames_draft = os.listdir(path)

    filenames = []                    # 골라낸 파일넣을 리스트 생성
    for i in filenames_draft:
        if '.xlsx' in i:              # 파일확장자 리스트 골라내기
            if '2D Text' in i:
                continue
            else:
                filenames.append(i)
    filenames.sort()                  # 리스트 abc순 정렬



    """ 엑셀파일생성 및 첫 시트 만들기 """
    df = pd.read_excel(f'{path}\\{filenames[0]}')

    p = re.compile("E\d{2}")
    m = p.search(filenames[0])
    wsname = m.group()

    # wsname = (filenames[0])[5:8]
    with pd.ExcelWriter(f'{path}\\{final_name}', mode='w', engine='openpyxl') as writer:
        df.to_excel(writer,sheet_name=wsname, index=False)


    """ 두번째시트부터 추가하기 """
    for file in filenames:
        p = re.compile("E\d{2}")
        m = p.search(file)
        wsname = m.group()
        if file != filenames[0]:                    # 첫번째 파일이 아니면 실행해라!
            df = pd.read_excel(f'{path}\\{file}')     # df생성
            # wsname = file[5:8]                      # 파일의 5:8 index값을 name으로 지정
            with pd.ExcelWriter(f'{path}\\{final_name}', mode='a', engine='openpyxl') as writer:
                df.to_excel(writer,sheet_name=wsname, index=False)


    wb = load_workbook(f'{path}\\{final_name}')
    sheet = wb.active
    wsnames = wb.get_sheet_names()      # 모든 시트명 가져오기
    for i,wsname in enumerate(wsnames):
        sheet = wb[wsname]              # 해당시트명의 시트 열기
        hearder=[]
        for seq, name in enumerate(range(1, sheet.max_column + 1)):
            header = sheet.cell(row=1, column=seq+1)
            header.font = Font(color="00000000", bold=True, size = 14)
            header.alignment = Alignment(horizontal="center", vertical="center")
        # sheet['A1'] = 'en-Origin'
        # sheet['A1'].font = Font(bold=True)
        # sheet['B1'] = 'en-Modity'
        # sheet['B1'].font = Font(bold=True)
        # sheet['C1'] = 'fr-CA'
        # sheet['C1'].font = Font(bold=True)
        sheet.column_dimensions["A"].width = 70
        sheet.column_dimensions["B"].width = 70
        sheet.column_dimensions["C"].width = 70
        for x in range(1, sheet.max_row + 1):
            sheet[f'A{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
            sheet[f'B{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
            sheet[f'C{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
        wb.save(f'{path}\\{final_name}')
    for file in filenames:              # 각각 애피소드 엑셀 제거 (E01.xlsx, E02.xlsx, ...)
        os.remove(f'{path}\\{file}')
    print(lang+' 완료')


'''
각 폴더별로 작업진행 코드
'''
for (dirpath, dirnames, filenames) in os.walk(path):   #정리할 파일 위치
    # print(dirpath)
    # print(filenames)
    p = re.compile("[a-zA-Z]{2}-[a-zA-Z0-9_]{2}")    # re.compile('..-[a-zA-Z]{2}')
    m = p.search(dirpath)         # ex)  <re.Match object; span=(0, 5), match='ar-AE'>
    try:    
        lang = m.group()                  # ex)  ar-AE
    except:
        continue

    if lang == 'es-41':
        lang = 'es-419'
    elif lang == 'uz-La':
        lang = 'uz-Latn-UZ'
    elif lang == 'az-La':
        lang = 'az-Latn-AZ'
    elif lang == 'bs-La':
        lang = 'bs-Latn-BA'
    elif lang == 'sr-La':
        lang = 'sr-Latn-RS'
    else:
        lang = lang

    # print(lang)                 # ru-RU
    # print(dirpath)              # C:\Users\USER\Documents\Export Files\ru-RU
    # print(filenames)            # ['(SRT)E05.srt', '(SRT)E10.srt', '(SRT)E17.srt', '(SSML)E05.txt', '(SSML)E10.txt', '(SSML)E17.txt', '2D Text_All.xlsx']

    # path를 dirpath로
    path_en = "D:\\Program Files\\Workspace\\00_Original\\"
    a1(dirpath,dirpath,path_en) # 각 함수 실행 a1
    a2(dirpath,lang)                 # 각 함수 실행 a2

