import os
import openpyxl

"""
엑셀통합파일의 각 시트(A열 영어원본,B열 영어수전본,C열 프랑스어)를, 영어수정본과 프랑스어 버전으로 srt파일, ssml파일로 각각 생성하기
"""

def createFolder(directory):
        try:
            if not os.path.exists(directory):
                os.makedirs(directory)
        except OSError:
            print ('Error: Creating directory. ' +  directory)

path= 'D:\\Program Files\\Workspace\\00_Default\\'
path2= 'D:\\Program Files\\Workspace\\00_Create\\'

excelfile_name_script = '(REF)Script_all_LGECI_0128.xlsx'
excelfile_name_2D = '(REF)_(US_TEXT)_2D_Text_LGECI_0128.xlsx'

path11 = 'D:\\Program Files\\Workspace\\00_Default\\US_1\\' # 원본파일srt
path12 = 'D:\\Program Files\\Workspace\\00_Default\\US_2\\' # 원본파일ssml

path13 = 'D:\\Program Files\\Workspace\\00_Create\\en-CA_1_srt\\'     # 변형파일srt
path14 = 'D:\\Program Files\\Workspace\\00_Create\\en-CA_2_ssml\\'    # 뱐형파일ssml

path15 = 'D:\\Program Files\\Workspace\\00_Create\\fr-CA_1_srt\\'     # 변형파일srt
path16 = 'D:\\Program Files\\Workspace\\00_Create\\fr-CA_2_ssml\\'    # 뱐형파일ssml

path17 = 'D:\\Program Files\\Workspace\\00_Create\\en-CA_3_text\\'    # 2D text en-CA
path18 = 'D:\\Program Files\\Workspace\\00_Create\\fr-CA_3_text\\'    # 2D text fr-CA

createFolder(path13)
createFolder(path14)
createFolder(path15)
createFolder(path16)
createFolder(path17)
createFolder(path18)

filenames_draft = os.listdir(path11)
filenames_path11 = []                    # 골라낸 파일넣을 리스트 생성
for i in filenames_draft:
    if '.srt' in i:              # 파일확장자 리스트 골라내기
        filenames_path11.append(i)
filenames_path11.sort()                  # 리스트 abc순 정렬

filenames_draft = os.listdir(path12)
filenames_path12 = []                    # 골라낸 파일넣을 리스트 생성
for i in filenames_draft:
    if '.txt' in i:              # 파일확장자 리스트 골라내기
        filenames_path12.append(i)
filenames_path12.sort()                  # 리스트 abc순 정렬


""" 원본srt,ssml불러오기 & 엑셀파일 열기 & replace(A1,B1)... """


wb = openpyxl.load_workbook(f'{path2}{excelfile_name_script}')    # replace(소스,타겟) data를 불러올 엘셀파일
get_sheet_names = wb.get_sheet_names()          # 시트이름들 리스트
print("\n")

"""1 en-CA SRT """
for k,filename in enumerate(filenames_path11):
    with open(f'{path11}{filenames_path11[k]}', 'r', encoding='utf-8') as file:     # 원본파일경로, 파일명
        content = file.read()
        sheet = wb[get_sheet_names[k]]
        for x in range(2, sheet.max_row + 1):
            A_value = sheet[f'A{x}'].value      # 소스 언어 열(A)
            B_value = sheet[f'B{x}'].value      # 변경될 언어 열(B)
            A_value = str(A_value).strip()
            B_value = str(B_value).strip()
            # print(A_value)
            # print(B_value)
            # print(sheet[f'A{x}'])        
            try:
                content = content.replace(A_value, B_value)
            except TypeError:
                if B_value == None:
                    content = content.replace(A_value, '')
    with open(f'{path13}[enCA]{filename}', "w", encoding='UTF-8') as file:           # 생성될 파일경로, 생성될파일명
            file.write(content)

"""2 en-CA SSML """
for k,filename in enumerate(filenames_path12):                                      
    with open(f'{path12}{filenames_path12[k]}', 'r', encoding='utf-8') as file:# 원본파일경로, 파일명
        content = file.read()
        sheet = wb[get_sheet_names[k]]
        for x in range(2, sheet.max_row + 1):
            A_value = sheet[f'A{x}'].value      # 소스 언어 열(A)
            B_value = sheet[f'B{x}'].value      # 변경될 언어 열(B)
            A_value = str(A_value).strip()
            B_value = str(B_value).strip()
            # print(A_value)
            # print(B_value)
            if not A_value in content:          # 검수용 코드 (바꾸려는 문장이, 대상파일 안에 없으면 프린트)
                print(A_value)                  # 검수용 코드 (바꾸려는 문장이, 대상파일 안에 없으면 프린트)
            try:
                content = content.replace(A_value, B_value)
            except TypeError:
                if B_value == None:
                    content = content.replace(A_value, '')
    with open(f'{path14}[enCA]{filename}', "w", encoding='UTF-8') as file:           # 생성될 파일경로, 생성될파일명
            file.write(content)

"""3 fr-CA SRT """
for k,filename in enumerate(filenames_path11):
    with open(f'{path11}{filenames_path11[k]}', 'r', encoding='utf-8') as file:# 원본파일경로, 파일명
        content = file.read()
        sheet = wb[get_sheet_names[k]]
        for x in range(2, sheet.max_row + 1):
            A_value = sheet[f'A{x}'].value      # 소스 언어 열(A)
            B_value = sheet[f'C{x}'].value      # 변경될 언어 열(C)
            A_value = str(A_value).strip()
            B_value = str(B_value).strip()
            # print(A_value)
            # print(B_value)
            try:
                content = content.replace(A_value, B_value)
            except TypeError:
                if B_value == None:
                    content = content.replace(A_value, '')
    with open(f'{path15}[frCA]{filename}', "w", encoding='UTF-8') as file:           # 생성될 파일경로, 생성될파일명
            file.write(content)

"""4 fr-CA SSML """
for k,filename in enumerate(filenames_path12):                                      
    with open(f'{path12}{filenames_path12[k]}', 'r', encoding='utf-8') as file:# 원본파일경로, 파일명
        content = file.read()
        sheet = wb[get_sheet_names[k]]
        for x in range(2, sheet.max_row + 1):
            A_value = sheet[f'A{x}'].value
            B_value = sheet[f'C{x}'].value      # 변경될 언어 열(C)
            A_value = str(A_value).strip()
            B_value = str(B_value).strip()            
            # print(A_value)
            # print(B_value)
            try:
                content = content.replace(A_value, B_value)
            except TypeError:
                if B_value == None:
                    content = content.replace(A_value, '')
    with open(f'{path16}[frCA]{filename}', "w", encoding='UTF-8') as file:           # 생성될 파일경로, 생성될파일명
            file.write(content)



print("\n")
""" 2D Text """
wb2 = openpyxl.load_workbook(f'{path2}{excelfile_name_2D}')    # replace(소스,타겟) data를 불러올 엘셀파일
get_sheet_names2 = wb2.get_sheet_names()          # 시트이름들 리스트

"""enCA 2D"""
for i,sheet_name in enumerate(get_sheet_names2):
    sheet = wb2[get_sheet_names2[i]]
    sheet.delete_rows(1)
    sheet.delete_cols(3)
    sheet.delete_cols(1)
wb2.save(f"{path17}[enCA].xlsx")

"""frCA 2D"""
wb2 = openpyxl.load_workbook(f'{path2}{excelfile_name_2D}')    # replace(소스,타겟) data를 불러올 엘셀파일
get_sheet_names2 = wb2.get_sheet_names()          # 시트이름들 리스트
for i,sheet_name in enumerate(get_sheet_names2):
    sheet = wb2[get_sheet_names2[i]]
    sheet.delete_rows(1)
    sheet.delete_cols(2)
    sheet.delete_cols(1)
wb2.save(f"{path18}[frCA].xlsx")
