import pandas as pd
import os
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook

Create_path = 'D:\\Program Files\\Workspace\\00_Create\\'
en_path = 'D:\\Program Files\\Workspace\\00_Default\\'
fr_path = 'D:\\Program Files\\Workspace\\00_Default_2\\'
en_filenames = os.listdir(en_path)
fr_filenames = os.listdir(fr_path)

en_txt_filenames = []                    # 골라낸 파일 리스트
for i in en_filenames:
    if '.txt' in i:              # 'txt' or 'srt'가 포함되어있는 리스트 골라내기
        en_txt_filenames.append(i)
en_txt_filenames.sort()

fr_txt_filenames = []
for i in fr_filenames:
    if '.txt' in i:              # 'txt' or 'srt'가 포함되어있는 리스트 골라내기
        fr_txt_filenames.append(i)
fr_txt_filenames.sort()


for n, en_file in enumerate(en_txt_filenames):

    en_line_list = []
    with open(f'{en_path}{en_txt_filenames[n]}','r',encoding = 'utf8') as File_Source:
        data = File_Source.readlines()
        for l in data:
            en_line_list.append(l.strip())

    fr_line_list = []
    with open(f'{fr_path}{fr_txt_filenames[n]}','r',encoding = 'utf8') as File_Source:
        data = File_Source.readlines()
        for l in data:
            fr_line_list.append(l.strip())


    """ 첫열 DF생성 """
    First_Data = [{'en-US':en_line_list[0]}]   # DF작성     # 첫 header 'en-US'
    df = pd.DataFrame(First_Data)              # DF생성
    i=1
    while i < 1000:
        try:
            df.loc[i]=[en_line_list[i]]                            # DF에 행추가
            i+=1
        except:
            break
    """ 두번째열 DF추가 """
    df['fr-CA'] = fr_line_list                              # 두번째 header 'fr-CA'
    # print(df)

    df.to_excel(f'{Create_path}test_{en_file[13:16]}.xlsx',sheet_name=f'{en_file[13:16]}', header = True, index = False)

    wb = load_workbook(f"{Create_path}test_{en_file[13:16]}.xlsx")
    ws = wb.active
    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 70
    for x in range(1, ws.max_row + 1):
    # for y in range(1, ws.max_column + 1):
        ws[f'A{x}'].alignment = Alignment(wrap_text = True)
        ws[f'B{x}'].alignment = Alignment(wrap_text = True)

    wb.save(f'{Create_path}test_{en_file[13:16]}.xlsx')

