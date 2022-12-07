import os
import re
import shutil
from openpyxl import load_workbook
import datetime


path = "C:\\Users\\USER\\Documents\\Export Files"
file2 = "C:\\Users\\USER\\Documents\\Export Files\\2D Text_All.xlsx"
dnow = datetime.datetime.now()

def input_2d_en(file,file2):        # excel 2Dtext 비교할수있도록 2열에 영어원본문장 삽입

    # 파일오픈
    wb = load_workbook(file)
    wb2 = load_workbook(file2)
    wsnames_1 = wb.get_sheet_names()      # 모든 시트명 가져오기
    wsnames_2 = wb2.get_sheet_names()      # 모든 시트명 가져오기
    for index_1, wsname_1 in enumerate(wsnames_1):
        print('////////// '+wsname_1+' //////////')
        sheet = wb[wsname_1]              # 해당시트명의 시트 열기
        sheet2 = wb2[wsname_1]              # 해당시트명의 시트 열기
        ## data 가져오기 ##
        for x in range(1, sheet.max_row + 1):
            # print(sheet2.cell(row=x, column=1).value)
            data = sheet2.cell(row=x, column=1).value       # 엑셀2의 sheet2의 x행 1열 data가져오기
            match = f"=MATCH(A{x},B{x},0)"
            sheet.cell(row=x, column=2, value=data)         # 엑셀1의 x행2열에 가져온data 삽입
            sheet.cell(row=x, column=3, value=match)        # 엑셀1의 x행3열에 match함수 삽입

    wb.save(file)


filename_list = []                       # 파일명 리스트
for (dirpath, dirnames, filenames) in os.walk(path):   #정리할 파일 위치
    # print(dirpath)
    # print(filenames)
    p = re.compile("[a-zA-Z]{2}-[a-zA-Z0-9_]{2}")    # re.compile('..-[a-zA-Z]{2}')
    m = p.search(dirpath)         # ex)  <re.Match object; span=(0, 5), match='ar-AE'>
    try:    
        lang = m.group()                  # ex)  ar-AE
    except:
        continue

    if lang == 'es-41':
        lang = 'es-419'
    elif lang == 'uz-La':
        lang = 'uz-Latn-UZ'
    elif lang == 'az-La':
        lang = 'az-Latn-AZ'
    elif lang == 'bs-La':
        lang = 'bs-Latn-BA'
    elif lang == 'sr-La':
        lang = 'sr-Latn-RS'
    else:
        lang = lang

    # print(lang)                 # ru-RU
    # print(dirpath)              # C:\Users\USER\Documents\Export Files\ru-RU
    # print(filenames)            # ['(SRT)E05.srt', '(SRT)E10.srt', '(SRT)E17.srt', '(SSML)E05.txt', '(SSML)E10.txt', '(SSML)E17.txt', '2D Text_All.xlsx']

    '''엑셀파일명 변경'''
    for i in filenames:
        if '.xlsx' in i:              # 'xlsx'가 포함되어있는 리스트 골라내기
            excel_name = i
    excel_rename = dirpath+"\\"+"BF_VB_2D Text_("+lang+")_"+dnow.strftime("%y%m%d")+".xlsx" # 파일명변경 # 제품군_제품명_2D Text_(언어코드)_날짜.xlsx
    os.rename(dirpath+"\\"+excel_name, excel_rename)                                        # 파일명변경 # 제품군_제품명_2D Text_(언어코드)_날짜.xlsx
    print(excel_rename)

    '''srt파일명 변경'''
    for i in filenames:
        if '.srt' in i:              # 'xlsx'가 포함되어있는 리스트 골라내기
            srt_name = i                             # (SRT)E05.srt
            m = re.compile("E\d{2}").search(i)       # <re.Match object; span=(12, 15), match='E05'>
            episode = m.group()                      # E05
            srt_rename = dirpath+"\\"+"(SRT)BF_VB_("+lang+")_"+episode+"_"+dnow.strftime("%y%m%d")+".srt"    # 제품군_제품명_ 에피소드No_Title _(언어코드)_날짜.srt
            os.rename(dirpath+"\\"+srt_name, srt_rename)                                                # 제품군_제품명_ 에피소드No_Title _(언어코드)_날짜.srt
            print(srt_rename)

    '''SSML파일명 변경'''
    for i in filenames:
        if '.txt' in i:              # 'xlsx'가 포함되어있는 리스트 골라내기
            ssml_name = i                             # (SRT)E05.srt
            m = re.compile("E\d{2}").search(i)       # <re.Match object; span=(12, 15), match='E05'>
            episode = m.group()                      # E05
            ssml_rename = dirpath+"\\"+"BF_VB_("+lang+")_"+episode+"_"+dnow.strftime("%y%m%d")+".txt"    # 제품군_제품명_ 에피소드No_Title _(언어코드)_날짜.srt
            os.rename(dirpath+"\\"+ssml_name, ssml_rename)                                                # 제품군_제품명_ 에피소드No_Title _(언어코드)_날짜.srt
            print(ssml_rename)

    '''엑셀파일 B열에 영어 업로드'''
    input_2d_en(excel_rename, file2)
