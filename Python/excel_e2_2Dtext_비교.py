import os
import pandas as pd
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
import re



path = 'D:\\Program Files\\Workspace\\00_Create\\'
file = 'D:\\Program Files\\Workspace\\00_Create\\(lv-LV)RF_(UK_TEXT)_220303.xlsx'
file2 = 'D:\\Program Files\\Workspace\\00_Create\\RF_(UK_TEXT)_220303.xlsx'


# 파일오픈
wb = load_workbook(file)
wb2 = load_workbook(file2)
wsnames_1 = wb.get_sheet_names()      # 모든 시트명 가져오기
wsnames_2 = wb2.get_sheet_names()      # 모든 시트명 가져오기
for index_1, wsname_1 in enumerate(wsnames_1):
    print('////////// '+wsname_1+' //////////')
    sheet = wb[wsname_1]              # 해당시트명의 시트 열기
    sheet2 = wb2[wsname_1]              # 해당시트명의 시트 열기
    ## data 가져오기 ##
    for x in range(1, sheet.max_row + 1):
        # print(sheet2.cell(row=x, column=1).value)
        data = sheet2.cell(row=x, column=1).value
        match = f"=MATCH(A{x},B{x},0)"
        sheet.cell(row=x, column=2, value=data)
        sheet.cell(row=x, column=3, value=match)

wb.save(file)