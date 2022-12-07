import os
import glob
import re
import os.path
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment


path_lang = "D:\\Program Files\\Workspace\\00_Default\\"
path_excel = "D:\\Program Files\\Workspace\\00_Create\\"
path_en = "D:\\Program Files\\Workspace\\00_Original\\"

###### <파일명 확장자 변경> #############################
files_lang = glob.glob(f'{path_lang}*.*')
files_en = glob.glob(f'{path_en}*.*')

for name in files_lang:

    if not os.path.isdir(name):  # 디렉토리는 포함 X
        filename1 = os.path.splitext(name)  # 확장자와 파일명 구분
        # print('filename1:',filename1[0],"//", filename1[1]) # 확장자와 파일명 구분 확인
        try:
            os.rename(name, filename1[0] + '.txt')  # ssml을 없앤 파일명에 txt 확장자 붙이기
        except:
            pass

for name in files_en:

    if not os.path.isdir(name):  # 디렉토리는 포함 X
        filename1_original = os.path.splitext(name)  # 확장자와 파일명 구분
        # print('filename1:',filename1[0],"//", filename1[1]) # 확장자와 파일명 구분 확인
        try:
            # ssml을 없앤 파일명에 txt 확장자 붙이기
            os.rename(name, filename1_original[0] + '.txt')
        except:
            pass
########################################################


##### <TXT파일로 확장자 변경후 파일 list 만들기> #########
filename_lang = []                  # 파일명 리스트
filename_lang_draft = os.listdir(path_lang)
for i in filename_lang_draft:
    if '.txt' in i:              # 'txt' 가 포함되어있는 리스트 골라내기
        filename_lang.append(i)

filename_en = []
filename_en_draft = os.listdir(path_en)
for o in filename_en_draft:
    if '.txt' in o:
        filename_en.append(o)


''' 엑셀파일 생성 '''

for n, filename in enumerate(filename_lang):

    en_line_list = []
    with open(f'{path_en}{filename_en[n]}','r',encoding = 'utf8') as File_Source:
        data = File_Source.readlines()
        data = list(map(lambda s: s.strip(), data))
        for l in data:
            en_line_list.append(l.strip())

    bb_line_list = []
    with open(f'{path_lang}{filename_lang[n]}','r',encoding = 'utf8') as File_Source:
        data = File_Source.readlines()
        data = list(map(lambda s: s.strip(), data))
        for l in data:
            bb_line_list.append(l.strip())

    dic = {
        'en': en_line_list,
        'Target language': bb_line_list,
        'Feedback': ''
    }

    df = pd.DataFrame.from_dict(dic, orient='index')
    df = df.transpose()     # DataFrame 행열 전환


    p = re.compile("E\d{2}")
    m = p.search(filename)
    episode = m.group()
    newfile_name = os.path.splitext(filename)[0]   # 확장자와 파일명 구분

    """ DataFrame을 엑셀로 뽑기 """
    df.to_excel(f'{path_excel}{newfile_name}.xlsx',sheet_name=f'{episode}', header = True, index = False)

    wb = load_workbook(f'{path_excel}{newfile_name}.xlsx')
    ws = wb.active
    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 70
    for x in range(1, ws.max_row + 1):
    # for y in range(1, ws.max_column + 1):
        ws[f'A{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
        ws[f'B{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
        ws[f'C{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
    wb.save(f'{path_excel}{newfile_name}.xlsx')
    print(episode,'완료')

