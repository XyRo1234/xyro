import tkinter as tk
from tkinter import PhotoImage, Button, Label, StringVar, ttk
import os
import os.path
from datetime import datetime
import re
import glob
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from chardet import detect
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import pyautogui


class App:

    global path1, path2, path3, path4, user_id, password

    # user_id, password = StringVar(), StringVar()
    
    def createFolder(directory):
        try:
            if not os.path.exists(directory):
                os.makedirs(directory)
        except OSError:
            print ('Error: Creating directory. ' +  directory)

    path1 = "D:\\Video_Manual\\1.Original\\"
    path2 = "D:\\Video_Manual\\2.Create\\"
    path3 = "D:\\Video_Manual\\GMDS\\"
    path4 = "D:\\Program Files\\Workspace\\picture\\"   
 

    createFolder(f'{path1}')
    createFolder(f'{path2}')
    createFolder(f'{path3}')


    def __init__(self):

        window = tk.Tk()
        window.geometry("490x680+300+150")  # 창 사이즈 + X위치 + Y위치
        window.title('어떤 도움을 드릴까요?                    (Build by KYS & JWS)')
        window.resizable(False, False)  # 창 X,Y 방향 사이즈 조절 막기

        buttonSize = (80, 70)

        photo1 = PhotoImage(file=f'{path4}exchange.png')
        photo2 = PhotoImage(file=f'{path4}SRT1.png')
        photo3 = PhotoImage(file=f'{path4}SSML1.png')
        photo4 = PhotoImage(file=f'{path4}TXT.png')
        photo5 = PhotoImage(file=f'{path4}XLS1.png')
        photo6 = PhotoImage(file=f'{path4}XLS2.png')
        photo7 = PhotoImage(file=f'{path4}topic1.png')
        photo8 = PhotoImage(file=f'{path4}para2.png')

        message1 = "파일 내의 특정 단어를 치환합니다.\n D:\\Video_Manual\\1.Original\\ 폴더에 파일을 넣으세요.\n ...\\2.Create\\ 폴더에 Changed_(파일명)으로 파일이 만들어집니다."
        message2 = "SRT 파일 내의 원문을 추출합니다.\n D:\\Video_Manual\\2.Create\\ 폴더에 파일을 넣으세요.\n 같은 폴더에 NEW_(파일명)으로 파일이 만들어집니다."
        message3 = "SSML 파일 내의 원문을 추출합니다.\n D:\\Video_Manual\\2.Create\\ 폴더에 파일을 넣으세요.\n 같은 폴더에 파일을 넣으세요.\n ...\\00_Create\\ 폴더에 NEW_(파일명)으로 파일이 만들어집니다."
        message4 = "파일 확장자를 txt로 변경합니다.\n D:\\Video_Manual\\2.Create\\ 폴더에 파일을 넣으세요."
        message5 = "TXT 파일을 엑셀 파일로 신규 생성해 줍니다.\n D:\\Video_Manual\\2.Create\\ 폴더에 파일을 넣으세요."
        message6 = "2D Text 엑셀 파일에 Header를 추가합니다.\n D:\\Video_Manual\\1.Original\\ 폴더에 파일을 넣으세요.\n ...\\2.Create\\ 폴더에 수정된 파일이 만들어집니다."
        message7 = "GMDS에 Topic을 자동 생성합니다.\n D:\\Video_Manual\\GMDS\\ \n 위 폴더에 작업 양식을 기입한 엑셀파일을 넣으세요."
        message8 = "GMDS Topic 내용을 자동 입력합니다.\n D:\\Video_Manual\\GMDS\\ \n 위 폴더에 작업 양식을 기입한 엑셀파일을 넣으세요."

        btn1 = Button(window, image=photo1,height=buttonSize[0], width=buttonSize[1], command=self.Check_word)
        btn1.grid(row=1, column=1)        # btn1.place(x=100,y=100) # x,y position으로 위치시키는 방법은 주석과 같이
        exp1 = Label(window, text=message1, font='Helvetica 10 bold')
        exp1.grid(row=1, column=2)
        
        btn2 = Button(window, image=photo2,height=buttonSize[0], width=buttonSize[1], command=self.SRT_Clear)
        btn2.grid(row=2, column=1)
        exp2 = Label(window, text=message2, font='Helvetica 10 bold')
        exp2.grid(row=2, column=2)
        
        btn3 = Button(window, image=photo3,height=buttonSize[0], width=buttonSize[1], command=self.SSML_Clear)
        btn3.grid(row=3, column=1)
        exp3 = Label(window, text=message3, font='Helvetica 10 bold')
        exp3.grid(row=3, column=2)
        
        btn4 = Button(window, image=photo4,height=buttonSize[0], width=buttonSize[1], command=self.change_TXT)
        btn4.grid(row=4, column=1)
        exp4 = Label(window, text=message4, font='Helvetica 10 bold')
        exp4.grid(row=4, column=2)
        
        btn5 = Button(window, image=photo5,height=buttonSize[0], width=buttonSize[1], command=self.change_Excel)
        btn5.grid(row=5, column=1)
        exp5 = Label(window, text=message5, font='Helvetica 10 bold')
        exp5.grid(row=5, column=2)
        
        btn6 = Button(window, image=photo6,height=buttonSize[0], width=buttonSize[1], command=self.TwoD_text)
        btn6.grid(row=6, column=1)
        exp6 = Label(window, text=message6, font='Helvetica 10 bold')
        exp6.grid(row=6, column=2)
        
        btn7 = Button(window, image=photo7,height=buttonSize[0], width=buttonSize[1], command=self.GMDS_Topic)
        btn7.grid(row=7, column=1)
        exp7 = Label(window, text=message7, font='Helvetica 10 bold')
        exp7.grid(row=7, column=2)

        btn8 = Button(window, image=photo8,height=buttonSize[0], width=buttonSize[1], command=self.GMDS_Data)
        btn8.grid(row=8, column=1)
        exp8 = Label(window, text=message8, font='Helvetica 10 bold')
        exp8.grid(row=8, column=2)

        window.mainloop()
      

    def Check_word(self):

        # 폴더내의 파일명을 리스트로 만들기
        filenames = os.listdir(path1)

        # 인코딩방식 체크
        def get_encoding_type(filepath):
            with open(filepath, "rb") as file:
                rawdata = file.read()
                return detect(rawdata)['encoding']
            return None

        # 파일내 단어수정
        for filename in filenames:
            # print(TXTSSML)

            src = []
            tar = []
            src.append("’")         # 기존내용      (right single quotation mark)
            tar.append("'")         # 바꾸려는 내용 (apostrophe)
            src.append("‘")         # 기존내용      (left single quotation mark)
            tar.append("'")         # 바꾸려는 내용 (apostrophe)
            src.append(" ")         # 기존내용      (NBSP)
            tar.append(" ")         # 바꾸려는 내용 (space)

            filepath = (f'{path1}{filename}')
            filepath2 = (f'{path2}Changed_{filename}')
            read_codec = get_encoding_type(filepath)

            if read_codec == "ISO-8859-1":
                with open(filepath, "r", encoding='ISO-8859-1') as file:           # r : read
                    content = file.read()
                    i = 0
                    while i <= 100:
                        try:
                            content = content.replace(src[i], tar[i])
                            i += 1
                        except IndexError:
                            break
                # w : write / a : append
                with open(filepath2, "a", encoding='UTF-8-SIG') as file:
                    file.write(content)

            elif read_codec == "UTF-8-SIG":
                with open(filepath, "r", encoding='utf-8') as file:           # r : read
                    content = file.read()
                    i = 0
                    while i <= 100:
                        try:
                            content = content.replace(src[i], tar[i])
                            i += 1
                        except IndexError:
                            break
                with open(filepath2, "a", encoding='UTF-8') as file:           # w : write / a : append
                    file.write(content)

            else:
                with open(filepath, "r", encoding='utf-8') as file:           # r : read
                    content = file.read()
                    i = 0
                    while i <= 100:
                        try:
                            content = content.replace(src[i], tar[i])
                            i += 1
                        except IndexError:
                            break
                # w : write / a : append
                with open(filepath2, "a", encoding='UTF-8-SIG') as file:
                    file.write(content)

    def SRT_Clear(self):
        
        '''폴더내 파일명 리스트 만들기'''

        # 폴더내의 파일명을 리스트로 만들기
        # filename = []                       # 파일명 리스트
        # for (dirpath, dirnames, filenames) in walk(r"D:\\Program Files\\Workspace\\00_Default"):  # 정리할 파일 위치
        #     filename.extend(filenames)
        #     break
        # print(filename)

        filename = os.listdir(path2)
        # 파일명리스트(filename)에서 txt or srt파일 골라내기.
        srt_file = []                       # 골라낸 파일 리스트
        for i in filename:
            if '.srt' in i:              # 'txt' or 'srt'가 포함되어있는 리스트 골라내기
                srt_file.append(i)

        ''' SRT 파일 불필요문구 제거 '''

        for file in srt_file:                                       # file : E01_How.srt
            with open(f'{path2}{file}', 'r', encoding='utf8') as fr:
                data = fr.readlines()
                for l in data:
                    line = l.strip()  # 양옆의 빈칸 삭제
                    try:
                        if line.isdigit():  # 숫자만 있는 것인지 확인
                            continue
                        # 문자열의 8자리를 datetime으로 변환을 하여 되는지 안되는지 확인
                        elif datetime.strptime(line[:8], '%H:%M:%S'):
                            continue
                    except ValueError:  # 정상적인 문자열이 for 문에서 except를 발생하여 정상 데이터를 저장함
                        if line != '':
                            if ord(line[0:1]) != 65279:
                                with open(f'{path2}New_{file}', 'a', encoding='utf8') as fl:
                                    fl.write(line.strip())
                                    fl.write('\n')

    def SSML_Clear(self):

        def extract_string(target):
            return re.sub('<[^>]*>', '', target)

        def Delete_BOM(target):
            target.replace(chr(65279), '')

        '''폴더내 파일명 리스트 만들기'''

        filenames = os.listdir(path2)


        # 파일명리스트(filename)에서 txt or srt파일 골라내기.
        txt_file = []                       # 골라낸 파일 리스트
        for i in filenames:
            if '.txt' in i:              # 'txt' or 'srt'가 포함되어있는 리스트 골라내기
                txt_file.append(i)
        # 파일명리스트(filename)에서 txt or srt파일 골라내기.
        ssml_file = []                       # 골라낸 파일 리스트
        for i in filenames:
            if '.ssml' in i:              # 'txt' or 'srt'가 포함되어있는 리스트 골라내기
                ssml_file.append(i)

        ''' txt 파일 불필요문구 제거 '''

        for file in txt_file:                                       # file : E01_How.txt
            with open(f'{path2}{file}', 'r', encoding='utf8') as fr:
                data = fr.readlines()
                for l in data:
                    p = extract_string(l.strip())  # < > 안의 데이터 모두 삭제
                    if p != '':  # 빈칸 삭제
                        if ord(p[0:1]) != 65279:
                            with open(f'{path2}New_{file}', 'a', encoding='utf8') as fl:
                                fl.write(p.strip())
                                fl.write('\n')

        ''' ssml파일 불필요문구 제거 '''
        for file in ssml_file:                                       # file : E01_How.txt
            with open(f'{path2}{file}', 'r', encoding='utf8') as fr:
                data = fr.readlines()
                for l in data:
                    p = extract_string(l.strip())  # < > 안의 데이터 모두 삭제
                    if p != '':  # 빈칸 삭제
                        if ord(p[0:1]) != 65279:
                            with open(f'{path2}New_{file}', 'a', encoding='utf8') as fl:
                                fl.write(p.strip())
                                fl.write('\n')

    def change_TXT(self):  # 파일명을 TXT로 변경합니다.

        # path1 = "D:\\Program Files\\Workspace\\00_Create\\"

        ###### <파일명 확장자 변경> #############################
        files = glob.glob(f'{path2}*.*')

        for name in files:

            if not os.path.isdir(name):  # 디렉토리는 포함 X
                filename1 = os.path.splitext(name)  # 확장자와 파일명 구분
                # print('filename1:',name)
                try:
                    # ssml을 없앤 파일명에 txt 확장자 붙이기
                    os.rename(name, filename1[0] + '.txt')
                except:
                    pass
        ########################################################

    def change_Excel(self):  # TXT파일을 엑셀 파일로 변경합니다.

        ##### <TXT파일로 확장자 변경후 파일 list 만들기> #########
        filename_draft = []
        filename = []                  # 파일명 리스트
        filename_draft = os.listdir(path2)
        for i in filename_draft:
            if '.txt' in i:              # 'txt' 가 포함되어있는 리스트 골라내기
                filename.append(i)

        ###### <엑셀 파일로 변경> #############################

        # print(filename, sep='\n')

        for name1 in filename:
            filename2 = os.path.splitext(name1)

            # 워크북(엑셀파일)을 새로 만듭니다.
            wb = openpyxl.Workbook()
            ws = wb.active  # 현재 시트 선택

        ###### Header 만들기 #########

            col_names = ['Original Script', 'Modified Script', 'French']
            hearder = []

            for seq, name in enumerate(col_names):
                header = ws.cell(row=1, column=seq+1, value=name)
                header.font = Font(color="000000FF", bold=True, size=14)
                header.alignment = Alignment(horizontal="center", vertical="center")

        ###### 본문 내용 넣기 #########

            # 각 파일의 내용을 줄단위로 읽어와서 list로 만든다.
            with open(f'{path2}{name1}', 'r', encoding='utf-8-sig') as fr:
                fr = fr.readlines()
                # relines() 읽은 값 뒤의 개행(\n) 없애주기
                fr = list(map(lambda s: s.strip(), fr))
                # fr 리스트의 맨 앞에 끼는 /ufeff는 utf-8-sig로 해결했으나 ""값으로 [0]번지에 들어가 있어 이를 지워줘야 엑셀 cell 하나를 잡아먹지 않는다.
                del fr[0]

            row_no = 2
            for n, script in enumerate(fr):
                ws.cell(row=row_no+n, column=1, value=script)

            ws.column_dimensions["A"].width = 70
            ws.column_dimensions["B"].width = 70
            ws.column_dimensions["C"].width = 70

            for x in range(2, ws.max_row + 1):
                ws[f'A{x}'].alignment = Alignment(
                    wrap_text=True, horizontal='left', vertical='center')
                ws[f'B{x}'].alignment = Alignment(
                    wrap_text=True, horizontal='left', vertical='center')
                ws[f'C{x}'].alignment = Alignment(
                    wrap_text=True, horizontal='left', vertical='center')
                ws.row_dimensions[x].height = 60

            wb.save(f'{path2}{filename2[0]}.xlsx')

    def TwoD_text(self):
        
        # path = 'D:\\Program Files\\Workspace\\others\\'
        filenames_draft = os.listdir(path1)

        filenames = []                    # 골라낸 파일넣을 리스트 생성
        for i in filenames_draft:
            if '.xlsx' in i:              # 파일확장자 리스트 골라내기
                filenames.append(i)
        filenames.sort()                  # 리스트 abc순 정렬

        for file in filenames:
            wb = load_workbook(f"{path1}{file}")
            ws = wb.active
            wsnames = wb.get_sheet_names()      # 모든 시트명 가져오기
            for i,wsname in enumerate(wsnames):
                sheet = wb[wsname]              # 해당시트명의 시트 열기
                sheet.insert_rows(1)            # 한줄 내리기
                sheet['A1'] = 'en-Origin'
                sheet['A1'].font = Font(bold=True)
                sheet['B1'] = 'en-Modify'
                sheet['B1'].font = Font(bold=True)
                sheet['C1'] = 'fr-CA'
                sheet['C1'].font = Font(bold=True)
                sheet.column_dimensions["A"].width = 50
                sheet.column_dimensions["B"].width = 50
                sheet.column_dimensions["C"].width = 50
                for x in range(1, sheet.max_row + 1):
                    sheet[f'A{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
                    sheet[f'B{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
                    sheet[f'C{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
            
            wb.save(f"{path2}{file}")


    def GMDS_Topic(self):  # GMDS Topic을 생성합니다.
        
        # tkinter 객체 생성
        win = tk.Toplevel()
        win.geometry("260x130+850+500")  # 창 사이즈 + X위치 + Y위치
        win.title('GMDS 로그인')
        win.resizable(False, False)  # 창 X,Y 방향 사이즈 조절 막기

        # 사용자 id와 password를 저장하는 변수 생성
        # global user_id,password
        user_id = StringVar()
        password = StringVar()
        
        def Go_Topic():
            url = 'http://gmds.lge.com/lgcms/topics'
            browser = webdriver.Ie('./IEDriverServer.exe')
            browser.get(url)
            browser.set_window_position(0,0)
            browser.set_window_size(1020,1080)
            browser.find_element_by_id("username").send_keys(user_id.get())
            # browser.find_element_by_id("password").send_keys(password.get())
            browser.find_element_by_class_name("login-button").click()
            time.sleep(1)

            GMDS_info_file = 'GMDS.xlsx'
            df_excel = pd.read_excel(f'{path3}{GMDS_info_file}', index_col=None)

            # US_titles = ['How to Use the Ice and Water Dispenser', 'How to Use the Ice Maker', 'How to Clean the Inside of the Refrigerator', 'How to Clean the Air Filter', 'How to Change the Water Filter (located inside the refrigerator)', 'Refrigerator_Not_Cooling', 'In Door Ice Maker Not Working', 'Craft Ice Maker Not Working', 'Water Is Not Dispensing', 'Water Tastes Bad', 'Refrigerator Door Not Closing', 'How to Find My Refrigerator Model Number', 'Choosing the Proper Installation Location', 'How to Remove and Assemble the Refrigerator Door', 'How_to_Connect_the_Water_Line', 'How to Align the Refrigerator Door', 'How to Level the Refrigerator']
            i = 0
            titles = []
            while i<100:
                try:
                    GMDS_title = str(df_excel.loc[i,'title'])
                    titles.append(GMDS_title)
                    i= i+1
                except KeyError:
                    break
            print(titles)

            Topic_No_list = []
            Topic_Title_list = []
            Region_list = []
            for title in titles:
                time.sleep(1)
                browser.find_element_by_xpath('/html/body/div[2]/div[2]/div[2]/button').click() # Create topic 클릭
                browser.find_element_by_xpath('//*[@id="title"]').send_keys(title)                                               # title 입력
                browser.find_element_by_xpath('/html/body/div[9]/div[2]/form/div[3]/div[1]/div/select/option[2]').click()           # Not Belong 클릭
                browser.find_element_by_xpath('//*[@id="product-id"]/option[35]').click()                                           # Product(*) : VideoGuidance 클릭
                browser.find_element_by_xpath('/html/body/div[9]/div[2]/form/div[5]/div[2]/div/select/optgroup/option[2]').click()  # Platform : SxS
                browser.find_element_by_xpath('/html/body/div[9]/div[2]/form/div[6]/div[1]/div/select/optgroup/option[1]').click()  # Region : GLOBAL
                # browser.find_element_by_xpath('/html/body/div[9]/div[2]/form/div[6]/div[1]/div/select/optgroup/option[2]').click()  # Region : NORTH AMERICA
                browser.find_element_by_xpath('/html/body/div[9]/div[2]/form/div[7]/div[1]/div/select/option[2]').click()           # Category : CC
                browser.find_element_by_xpath('/html/body/div[9]/div[2]/form/div[9]/button[3]').click()                             # Create 클릭
                time.sleep(1)
                Topic_No = browser.find_elements_by_class_name('topic-index-topic-info-open')[0].text       # Topic_No 정보가져오기
                Topic_No_list.append(Topic_No)
                Topic_Title = browser.find_elements_by_class_name('topic-index-topic-info-open')[2].text    # Topic_Title 정보가져오기
                Topic_Title_list.append(Topic_Title)
                Region = browser.find_elements_by_class_name('topic-index-topic-info-open')[6].text         # Region 정보가져오기
                Region_list.append(Region)

            # print(Topic_No_list)
            # print(Topic_Title_list)
            # print(Region_list)

            """ DataFrame 만들기 (Topic_No / Topic_Title / Region) """
            First_Data = [{'No':Topic_No_list[0], 'Topic Title':Topic_Title_list[0],'Region':Region_list[0]}]
            df = pd.DataFrame(First_Data)
            i=1
            while i < 100:
                try:
                    df.loc[i]=[Topic_No_list[i], Topic_Title_list[i],Region_list[i]]
                    i+=1
                except:
                    break
            print(df)

        # 엑셀파일 경로 지정
            GMDS_info_file = 'GMDS_Topic_Info.csv'                           # 엑셀파일명 지정
            df.to_csv(f'{path3}{GMDS_info_file}', header = True, index = True) 

       
        # id와 password, 그리고 확인 버튼의 UI를 만드는 부분
        ttk.Label(win, text = "Username : ").grid(row = 0, column = 0, padx = 10, pady = 10)
        ttk.Label(win, text = "Password : ").grid(row = 1, column = 0, padx = 10, pady = 10)
        ttk.Entry(win, textvariable = user_id).grid(row = 0, column = 1, padx = 10, pady = 10)
        ttk.Entry(win, textvariable = password, show='*').grid(row = 1, column = 1, padx = 10, pady = 10)
        ttk.Button(win, text = "Login", command = Go_Topic).grid(row = 2, column = 1, padx = 10, pady = 10)

             
        win.mainloop()

    def GMDS_Data(self):  # GMDS Topic을 생성합니다.
        
        # tkinter 객체 생성
        win = tk.Toplevel()
        win.geometry("260x130+850+500")  # 창 사이즈 + X위치 + Y위치
        win.title('GMDS 로그인')
        win.resizable(False, False)  # 창 X,Y 방향 사이즈 조절 막기

        # 사용자 id와 password를 저장하는 변수 생성
        # global user_id,password
        user_id = StringVar()
        password = StringVar()    
    
        def Go_Data():
            # """ line 리스트 만들기 """

            # path2 = "D:\\Program Files\\Workspace\\00_Create\\"
            # filenames = os.listdir(path2)
            # ## 파일명리스트(filename)에서 txt or srt파일 골라내기.
            # srt_file = []                       # 골라낸 파일 리스트
            # for i in filenames:
            #     if '.srt' in i:              # 'txt' or 'srt'가 포함되어있는 리스트 골라내기
            #         srt_file.append(i)
            # ## 파일명리스트(filename)에서 txt or srt파일 골라내기.
            # txt_file = []                       # 골라낸 파일 리스트
            # for i in filenames:
            #     if '.txt' in i:              # 'txt' or 'srt'가 포함되어있는 리스트 골라내기
            #         txt_file.append(i)
            # ## 파일명리스트(filename)에서 txt or srt파일 골라내기.
            # xml_file = []                       # 골라낸 파일 리스트
            # for i in filenames:
            #     if '.xml' in i:              # 'txt' or 'srt'가 포함되어있는 리스트 골라내기
            #         xml_file.append(i)

            browser = webdriver.Ie('./IEDriverServer.exe')
            url = f'http://gmds.lge.com/lgcms/topics'
            browser.get(url)
            browser.set_window_position(0,0)
            browser.set_window_size(1020,600)
            browser.find_element_by_id("username").send_keys(user_id.get())
            # browser.find_element_by_id("password").send_keys(password.get())
            browser.find_element_by_class_name("login-button").click()

            GMDS_info_file = 'GMDS.xlsx'                           # 엑셀파일 명
            df = pd.read_excel(f'{path3}{GMDS_info_file}', index_col=None)   # 엑셀 DataFrame 가져오기

            i=0
            while i<1000:
                try:
                    lines = []
                    No = int(df.loc[i,'No'])
                    # GMDS_episode = str(df.loc[i,'episode'])
                    # GMDS_title = str(df.loc[i,'title'])
                    GMDS_filename = str(df.loc[i,'filename'])               # 불러올 파일명(txt)
                    print(f'{No}:{GMDS_filename.rjust(40)}')
                    lines = []
                    with open(f'{path2}{GMDS_filename}','r',encoding = 'utf8') as fr:
                        data = fr.readlines()
                        for l in data:
                            if l != '':
                                lines.append(l)
                    i= i+1

                    """topic에 로그인 및 접속 """
                    url = f'http://gmds.lge.com/lgcms/Editor/src/topic?num=000{No}&r=1'       
                    url = f'http://gmds.lge.com/lgcms/Editor/src/topic?num=00069178&r=1'
                    browser.get(url)
                    browser.set_window_size(1020,600)
                    time.sleep(2)

                    """ Variablelist 추가 """
                    browser.find_element_by_id('treeToggle').click()
                    actionChains = ActionChains(browser)
                    target = browser.find_element_by_xpath('/html/body/div[3]/div[3]/div[2]/div[1]/div[2]').click() # topic 선택
                    actionChains.context_click(target).perform()    # 우클릭
                    browser.find_element_by_xpath('/html/body/div[30]/div[1]/div[6]').click()   # Insert element 0-4
                    browser.find_element_by_xpath('/html/body/div[30]/div[3]/div[3]').click()   # Insert to child 2-1
                    browser.find_element_by_xpath('/html/body/div[30]/div[5]/div[4]').click()   # variablelist 4-3

                    """ 마우스 이동 (x 좌표, y 좌표) """
                    # position = pyautogui.position()
                    # print(position.x)
                    # print(position.y)
                    pyautogui.moveTo(345,355)
                    pyautogui.click()

                    lines_modify = []
                    for l in lines:
                        l = l.strip()
                        if l != '':
                            lines_modify.append(l)

                    p = re.compile("^-")
                    for index, l in enumerate(lines_modify):
                        l = l.strip()
                        m = p.match(l)
                        if m:
                            # target = browser.find_element_by_xpath('/html/body/div[3]/div[3]/div[2]/div[1]/div[2]').click() # topic 선택
                            target = browser.find_elements_by_class_name('item')[0].click() # topic 선택 (이상동작이 추가 발생해서 한번 더 선택코드 삽입)
                            target = browser.find_elements_by_class_name('item')[0].click() # topic 선택 (위 이유로 한번 더 선택)                              
                            actionChains.context_click(target).perform()    # 우클릭
                            browser.find_element_by_xpath('/html/body/div[30]/div[1]/div[6]').click()   # Insert element 0-4
                            browser.find_element_by_xpath('/html/body/div[30]/div[3]/div[3]').click()   # Insert to child 2-1
                            browser.find_element_by_xpath('/html/body/div[30]/div[5]/div[4]').click()   # variablelist 4-3
                            l = (re.sub(p,'', l)).strip()        # 첫 - 없애기
                            actionChains.send_keys(l).perform()  # 소제목 타이핑
                            # actionChains.send_keys(Keys.ENTER).perform()  # Enter 타이핑
                            pyautogui.moveTo(345,475)
                            pyautogui.click()           # Para 클릭 (클릭후 왼쪽 Structure View에서 구조이동이 발생하면서 클릭이 풀리는 현상이 발생함)
                            time.sleep(1)               # 1초동안 구조이동시간 기다림
                            pyautogui.moveTo(345,475)
                            pyautogui.click()           # Para 재클릭 (구조이동이 완료된 후에 다시 클릭을 진행해줌)
                        else:
                            actionChains.send_keys(l).perform()  # line값 타이핑
                            if len(lines_modify)-1 == index:                # 마지막 index일때, 엔터를 치지 않음
                                continue
                            elif p.match(lines_modify[(index)+1]):          # 다음 index가 소제목(-)조건일 경우, 엔터를 치지 않음
                                continue
                            else:
                                actionChains.send_keys(Keys.ENTER).perform()  # Enter 타이핑

                    time.sleep(1)
                    
                    """ Save """
                    # browser.find_element_by_id('f').click()                     # File 탭 클릭
                    # browser.find_element_by_id('_mulcome_obj_0-1').click()      # Save 클릭
                
                except KeyError:
                    print("마지막 Topic 넘버를 불러왔었기에 종료합니다.")
                    break
                except ValueError:
                    print("마지막 Topic 넘버를 불러왔었기에 종료합니다.")
                    break       

        # id와 password, 그리고 확인 버튼의 UI를 만드는 부분
        ttk.Label(win, text = "Username : ").grid(row = 0, column = 0, padx = 10, pady = 10)
        ttk.Label(win, text = "Password : ").grid(row = 1, column = 0, padx = 10, pady = 10)
        ttk.Entry(win, textvariable = user_id).grid(row = 0, column = 1, padx = 10, pady = 10)
        ttk.Entry(win, textvariable = password, show='*').grid(row = 1, column = 1, padx = 10, pady = 10)
        ttk.Button(win, text = "Login", command = Go_Data).grid(row = 2, column = 1, padx = 10, pady = 10)

App()