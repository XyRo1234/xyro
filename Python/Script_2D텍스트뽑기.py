from tkinter import * # __all__
from tkinter import filedialog
import tkinter as tk
from tkinter import PhotoImage, Button, Label, StringVar, ttk, colorchooser
import tkinter.ttk as ttk
import tkinter.messagebox as msgbox
import os
import pandas as pd
from pyairtable import Table
import re
import numpy
from chardet import detect

from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook

path = 'D:\\Program Files\\Workspace\\airtable'
file1 = 'D:\Program Files\Workspace\\airtable\Script_E.05-Grid view.csv'
file2 = 'D:\Program Files\Workspace\\airtable\Script_E.10-Grid view.csv'
file3 = 'D:\Program Files\Workspace\\airtable\Script_E.15-Grid view.csv'
file4 = 'D:\Program Files\Workspace\\airtable\Script_E.17-Grid view.csv'
file5 = 'D:\Program Files\Workspace\\airtable\Script_E.21-Grid view.csv'
file6 = 'D:\Program Files\Workspace\\airtable\Script_E.22-Grid view.csv'

excelfiles = [file1,file2,file3,file4,file5,file6]

def start_1(excelfiles):
    #파일 불러오기
    for i,excelfile in enumerate(excelfiles):
        if '.csv' in excelfile:
            try:
                df = pd.read_csv(f'{excelfile}', encoding='CP949',  na_values=["?", "??", "N/A", ""], keep_default_na= False)  # 엑셀값 dataframe만들기
            except UnicodeDecodeError:
                df = pd.read_csv(f'{excelfile}', encoding='utf-8',  na_values=["?", "??", "N/A", ""], keep_default_na= False)  # 엑셀값 dataframe만들기
        else:
            df = pd.read_excel(f'{excelfile}')
        print(i)
        print(excelfile)

        header = df.columns.values.tolist()     # BOM값으로 인해 첫열 이름이 이상하게 변형되었을때, 열이름 변경하기
        if '癤풬o' in header:                   # BOM값으로 인해 첫열 이름이 이상하게 변형되었을때, 열이름 변경하기
            df.rename(columns={'癤풬o':'Ep_No'}, inplace=True) # 열이름 변경하기
        df = df.filter(items=['2D Text'])     # df필요한 열만 추출

        p = re.compile("\d{2}")
        m = p.search(excelfile)
        newfile_name = 'E'+m.group() # E01 등

        # xlsx_name = excelfile.split('.csv')[0] + '.xlsx'    # 파일명 확장자 csv -> xlsx로 바꾸기
        xlsx_name = path + '\\(en-GB)_2D Text_All.xlsx'

        """ 엑셀파일생성 및 첫 시트 만들기 """
        if excelfile == excelfiles[0]:                # 첫번째 파일이면 실행해라!
            with pd.ExcelWriter(f'{xlsx_name}', mode='w', engine='openpyxl') as writer: # 엑셀생성
                df.to_excel(writer,sheet_name=newfile_name, index=False)                # 엑셀생성
        else:                                         # 첫번째 파일이 아니면 실행해라!
            # wsname = file[5:8]                      # 파일의 5:8 index값을 name으로 지정
            with pd.ExcelWriter(f'{xlsx_name}', mode='a', engine='openpyxl') as writer: # 엑셀시트추가
                df.to_excel(writer,sheet_name=newfile_name, index=False)                # 엑셀시트추가
        
    wb = load_workbook(xlsx_name)
    sheet = wb.active
    wsnames = wb.get_sheet_names()      # 모든 시트명 가져오기
    for i,wsname in enumerate(wsnames):
        sheet = wb[wsname]              # 해당시트명의 시트 열기
        sheet.column_dimensions["A"].width = 70
        sheet.column_dimensions["B"].width = 70
        sheet.column_dimensions["C"].width = 70
        sheet.delete_rows(1)
        for x in range(1, sheet.max_row + 1):
            sheet[f'A{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
            sheet[f'B{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
            sheet[f'C{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
        wb.save(xlsx_name)






        # df.to_excel(f'D:\\Program Files\\Workspace\\airtable\\test_{newfile_name}.xlsx',sheet_name=f'{newfile_name}', header = True, index = False)

        # """뽑은엑셀에서 열너비 지정, 오른쪽가운데정렬, 자동줄바꿈 활성화 시켜서 저장"""
        # wb = load_workbook(f"D:\\Program Files\\Workspace\\airtable\\test_{newfile_name}.xlsx")
        # ws = wb.active
        # ws.column_dimensions["A"].width = 10
        # ws.column_dimensions["B"].width = 10
        # ws.column_dimensions["C"].width = 10
        # for x in range(1, ws.max_row + 1):
        # # for y in range(1, ws.max_column + 1):
        #     ws[f'A{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
        #     ws[f'B{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
        #     ws[f'C{x}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text = True)
        # wb.save(f'D:\\Program Files\\Workspace\\airtable\\test_{newfile_name}.xlsx')

        # print(newfile_name+'완료')








# """ 엑셀파일생성 및 첫 시트 만들기 """
# df = pd.read_excel(f'{path}{filenames[0]}')

# p = re.compile("E\d{2}")
# m = p.search(filenames[0])
# wsname = m.group()

# # wsname = (filenames[0])[5:8]
# with pd.ExcelWriter(f'{path}{final_name}', mode='w', engine='openpyxl') as writer:
#     df.to_excel(writer,sheet_name=wsname, index=False)


# """ 두번째시트부터 추가하기 """
# for file in filenames:
#     p = re.compile("E\d{2}")
#     m = p.search(file)
#     wsname = m.group()
#     if file != filenames[0]:                    # 첫번째 파일이 아니면 실행해라!
#         df = pd.read_excel(f'{path}{file}')     # df생성
#         # wsname = file[5:8]                      # 파일의 5:8 index값을 name으로 지정
#         with pd.ExcelWriter(f'{path}{final_name}', mode='a', engine='openpyxl') as writer:
#             df.to_excel(writer,sheet_name=wsname, index=False)






start_1(excelfiles)