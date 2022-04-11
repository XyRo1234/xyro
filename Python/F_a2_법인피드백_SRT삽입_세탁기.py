import os
from posixpath import split
import pandas as pd
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
import re
from datetime import datetime
import pandas as pd


path = 'D:\\영상콘텐츠\\세탁기\\언어Vari_220221_2차'

def line_list(path, Episode):
    '''폴더내 파일명 리스트 만들기'''
    ## 폴더내의 파일명을 리스트로 만들기
    filename = []                       # 파일명 리스트                     
    for (dirpath, dirnames, filenames) in os.walk(f"{path}\\1__subtitles"):   #정리할 파일 위치
        for a in filenames:
            filename.append(dirpath + '\\' + a)
    ## 파일명리스트(filename)에서 txt or srt파일 골라내기.
    srt_file = []                       # 골라낸 파일 리스트
    for i in filename:
        if '.srt' in i:              # 'txt' or 'srt'가 포함되어있는 리스트 골라내기
            srt_file.append(i)
    ''' SRT파일 불필요문구 제거 '''
    # file : E01_How.srt
    L_list = []
    for file in srt_file:
        if Episode in file:
            with open(file,'r',encoding = 'utf8') as fr:
                data = fr.readlines()
                for l in data:
                    line = l.strip() # 양옆의 빈칸 삭제
                    try:
                        if line.isdigit(): # 숫자만 있는 것인지 확인
                            continue
                        elif datetime.strptime(line[:8],'%H:%M:%S'): #문자열의 8자리를 datetime으로 변환을 하여 되는지 안되는지 확인
                            continue
                    except ValueError:  # 정상적인 문자열이 for 문에서 except를 발생하여 정상 데이터를 저장함
                        if line != '':
                            if ord(line[0:1]) != 65279:
                                L_list.append(line.strip())

    return L_list


Draft_path_file_list = []         # srt, ssml, 2D text 모두 경로+파일명 리스트로 만들기
for (path, dir, files) in os.walk(path):
    for a in files:
        Draft_path_file_list.append(path + '\\' + a)

path_file_list = []
for i in Draft_path_file_list:
    if not "2Dtext" in i:
        if '.xlsx' in i:              # 'xlsx'가 포함되어있는 리스트 골라내기
            print(i)
            path_file_list.append(i)


for index, excelname in enumerate(path_file_list):
    CC_path = excelname.split("\\LG")[0]
    number = index+1
    print("파일열기: ",excelname, "[" + str(number) + "/" + str(len(path_file_list)) + "]")
    wb = load_workbook(excelname)
    wsnames_1 = wb.get_sheet_names()      # 모든 시트명 가져오기
    for wsname_1 in wsnames_1:
        print('////////// '+wsname_1+' //////////')
        sheet = wb[wsname_1]              # 해당시트명의 시트 열기
        x = 30
        lines = line_list(CC_path, wsname_1)
        for line in lines:
            sheet.cell(row=x, column=5, value=line)     # data 입력
            x = x+1
            wb.save(excelname)


